"""
Merge Excel consumption data + Census centroids into final national county dataset.
"""
import json, os
import openpyxl

EXCEL_PATH = r'C:\Users\Danie\Downloads\Project Ultrasound - LIVE.xlsx'
CENTROIDS = r'C:\Users\Danie\Downloads\se-propane-market-map\data\county_centroids.json'
EXISTING = r'C:\Users\Danie\Downloads\se-propane-market-map\data\counties.json'
OUTPUT = r'C:\Users\Danie\Downloads\se-propane-market-map\data\counties_national.json'

STATE_ABBREV = {
    'Alabama': 'AL', 'Alaska': 'AK', 'Arizona': 'AZ', 'Arkansas': 'AR',
    'California': 'CA', 'Colorado': 'CO', 'Connecticut': 'CT', 'Delaware': 'DE',
    'District of Columbia': 'DC', 'Florida': 'FL', 'Georgia': 'GA', 'Hawaii': 'HI',
    'Idaho': 'ID', 'Illinois': 'IL', 'Indiana': 'IN', 'Iowa': 'IA',
    'Kansas': 'KS', 'Kentucky': 'KY', 'Louisiana': 'LA', 'Maine': 'ME',
    'Maryland': 'MD', 'Massachusetts': 'MA', 'Michigan': 'MI', 'Minnesota': 'MN',
    'Mississippi': 'MS', 'Missouri': 'MO', 'Montana': 'MT', 'Nebraska': 'NE',
    'Nevada': 'NV', 'New Hampshire': 'NH', 'New Jersey': 'NJ', 'New Mexico': 'NM',
    'New York': 'NY', 'North Carolina': 'NC', 'North Dakota': 'ND', 'Ohio': 'OH',
    'Oklahoma': 'OK', 'Oregon': 'OR', 'Pennsylvania': 'PA', 'Rhode Island': 'RI',
    'South Carolina': 'SC', 'South Dakota': 'SD', 'Tennessee': 'TN', 'Texas': 'TX',
    'Utah': 'UT', 'Vermont': 'VT', 'Virginia': 'VA', 'Washington': 'WA',
    'West Virginia': 'WV', 'Wisconsin': 'WI', 'Wyoming': 'WY'
}

def normalize_name(name):
    parts = name.rsplit(', ', 1)
    if len(parts) == 2:
        county, state = parts
        abbrev = STATE_ABBREV.get(state.strip(), state.strip())
        return f"{county.strip()}, {abbrev}"
    return name

# Load centroids
with open(CENTROIDS) as f:
    centroids = json.load(f)
print(f"Centroids: {len(centroids)}")

# Load existing SE data (as fallback for coords)
with open(EXISTING) as f:
    existing = {c['n']: c for c in json.load(f)}
print(f"Existing SE counties: {len(existing)}")

# Load Excel
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Consol']

national = []
matched_centroid = 0
matched_existing = 0
no_coords = 0
skipped = 0

for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    name = row[1]
    if not name:
        continue

    households = row[5]
    usage_per_hh = row[6]
    weighted_use = row[7]

    try:
        h = int(float(households)) if households else 0
        u = round(float(usage_per_hh), 1) if usage_per_hh else 0
        g = int(float(weighted_use)) if weighted_use else 0
    except (ValueError, TypeError):
        skipped += 1
        continue

    norm = normalize_name(name)

    entry = {'n': norm, 'h': h, 'u': u, 'g': g}

    # Try Census centroids first (most authoritative)
    if norm in centroids:
        entry['lat'] = centroids[norm]['lat']
        entry['lng'] = centroids[norm]['lng']
        entry['fips'] = centroids[norm]['fips']
        matched_centroid += 1
    elif norm in existing:
        entry['lat'] = existing[norm]['lat']
        entry['lng'] = existing[norm]['lng']
        matched_existing += 1
    else:
        # Try fuzzy match - strip "County", "Parish", etc.
        found = False
        for ckey, cval in centroids.items():
            if norm.lower() == ckey.lower():
                entry['lat'] = cval['lat']
                entry['lng'] = cval['lng']
                entry['fips'] = cval['fips']
                matched_centroid += 1
                found = True
                break
        if not found:
            entry['lat'] = None
            entry['lng'] = None
            no_coords += 1

    national.append(entry)

# Sort by total gallons descending
national.sort(key=lambda x: x['g'], reverse=True)

print(f"\n=== Results ===")
print(f"Total counties: {len(national)}")
print(f"Matched via Census centroids: {matched_centroid}")
print(f"Matched via existing SE data: {matched_existing}")
print(f"No coordinates: {no_coords}")
print(f"Skipped (bad data): {skipped}")

# Show counties without coords
if no_coords > 0:
    print(f"\n=== Counties Missing Coordinates ({no_coords}) ===")
    missing = [c for c in national if c['lat'] is None]
    for c in missing[:30]:
        print(f"  {c['n']}")

# Save
with open(OUTPUT, 'w') as f:
    json.dump(national, f, separators=(',', ':'))
size = os.path.getsize(OUTPUT)
print(f"\nSaved: {OUTPUT} ({size:,} bytes, {size/1024:.0f} KB)")

# State summary
states = {}
for c in national:
    parts = c['n'].rsplit(', ', 1)
    st = parts[1] if len(parts) == 2 else '?'
    if st not in states:
        states[st] = {'count': 0, 'gal': 0, 'coords': 0}
    states[st]['count'] += 1
    states[st]['gal'] += c['g']
    if c['lat']:
        states[st]['coords'] += 1

print(f"\n=== All States Summary ===")
for st in sorted(states.keys()):
    s = states[st]
    print(f"  {st}: {s['count']} counties, {s['gal']:>12,} gal, {s['coords']}/{s['count']} coords")
