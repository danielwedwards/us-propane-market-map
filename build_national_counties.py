"""
Build national county dataset from Excel workbook + existing SE county data.
Outputs: data/counties_national.json with all ~3,118 counties.
"""
import json, os, re
import openpyxl

EXCEL_PATH = r'C:\Users\Danie\Downloads\Project Ultrasound - LIVE.xlsx'
EXISTING_COUNTIES = r'C:\Users\Danie\Downloads\se-propane-market-map\data\counties.json'
OUTPUT = r'C:\Users\Danie\Downloads\se-propane-market-map\data\counties_national.json'

# ─── Step 1: Load existing SE county data (has lat/lng for 961 counties) ───
with open(EXISTING_COUNTIES) as f:
    existing = json.load(f)

existing_lookup = {}
for c in existing:
    existing_lookup[c['n']] = c
print(f"Existing SE counties with coordinates: {len(existing_lookup)}")

# ─── Step 2: Load Excel Consol sheet (all 3,118 counties) ───
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
ws = wb['Consol']

excel_counties = {}
for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
    name = row[1]
    if not name:
        continue
    households = row[5]
    usage_per_hh = row[6]
    weighted_use = row[7]
    try:
        h = int(float(households)) if households else 0
        u = round(float(usage_per_hh), 1) if usage_per_hh else 0
        g = int(float(weighted_use)) if weighted_use else 0
    except (ValueError, TypeError):
        continue
    excel_counties[name] = {'h': h, 'u': u, 'g': g}

print(f"Excel counties loaded: {len(excel_counties)}")

# ─── Step 3: Load Census household data for cross-reference ───
ws_census = wb['Census Data, by County']
census_hh = {}
for row in ws_census.iter_rows(min_row=2, max_row=ws_census.max_row, values_only=True):
    name = row[1]
    hh = row[2]
    if name and hh:
        try:
            census_hh[name] = int(float(hh))
        except:
            pass
print(f"Census household data: {len(census_hh)} counties")

# ─── Step 4: Build US county centroid lookup ───
# We'll use the us_zip_coords.py file which has coordinate data,
# or we'll need to geocode. First check what we have.

# Try to load county FIPS -> centroid mapping from a known source
# The US Census Gazetteer files have county centroids
# For now, let's see how many we can match from existing data

# State abbreviation mapping
STATE_ABBREV = {
    'Alabama': 'AL', 'Alaska': 'AK', 'Arizona': 'AZ', 'Arkansas': 'AR',
    'California': 'CA', 'Colorado': 'CO', 'Connecticut': 'CT', 'Delaware': 'DE',
    'District of Columbia': 'DC', 'Florida': 'FL', 'Georgia': 'GA', 'Hawaii': 'HI',
    'Idaho': 'ID', 'Illinois': 'IL', 'Indiana': 'IN', 'Iowa': 'IA',
    'Kansas': 'KS', 'Kentucky': 'KY', 'Louisiana': 'LA', 'Maine': 'ME',
    'Maryland': 'MD', 'Massachusetts': 'MA', 'Michigan': 'MI', 'Minnesota': 'MN',
    'Mississippi': 'MS', 'Missouri': 'MO', 'Montana': 'MT', 'Nebraska': 'NE',
    'Nevada': 'NV', 'New Hampshire': 'NH', 'New Jersey': 'NJ', 'New Mexico': 'NM',
    'New York': 'NY', 'North Carolina': 'NC', 'North Dakota': 'ND', 'Ohio': 'OH',
    'Oklahoma': 'OK', 'Oregon': 'OR', 'Pennsylvania': 'PA', 'Rhode Island': 'RI',
    'South Carolina': 'SC', 'South Dakota': 'SD', 'Tennessee': 'TN', 'Texas': 'TX',
    'Utah': 'UT', 'Vermont': 'VT', 'Virginia': 'VA', 'Washington': 'WA',
    'West Virginia': 'WV', 'Wisconsin': 'WI', 'Wyoming': 'WY'
}

# Normalize county names for matching: "County Name, State" -> abbreviated
def normalize_name(name):
    """Convert 'Autauga County, Alabama' to 'Autauga County, AL'"""
    parts = name.rsplit(', ', 1)
    if len(parts) == 2:
        county, state = parts
        abbrev = STATE_ABBREV.get(state.strip(), state.strip())
        return f"{county.strip()}, {abbrev}"
    return name

# Check how many Excel counties match existing SE data
matched = 0
unmatched_states = set()
for excel_name in excel_counties:
    norm = normalize_name(excel_name)
    if norm in existing_lookup:
        matched += 1
    else:
        parts = excel_name.rsplit(', ', 1)
        if len(parts) == 2:
            unmatched_states.add(parts[1].strip())

print(f"\nExcel counties matching existing SE data: {matched}/{len(excel_counties)}")
print(f"States needing coordinates: {len(unmatched_states)}")
print(f"Non-SE states: {sorted(unmatched_states - {'Alabama','Florida','Georgia','Kentucky','Louisiana','Mississippi','North Carolina','South Carolina','Tennessee','Virginia','Arkansas'})}")

# ─── Step 5: We need county centroids for the missing counties ───
# Let's download the US Census Gazetteer county centroids
# URL: https://www2.census.gov/geo/docs/maps-data/data/gazetteer/2023_Gazetteer/2023_Gaz_counties_national.txt
# For now, let's create a placeholder script that tells us what we need

# Check if we have a centroid file already
centroid_file = r'C:\Users\Danie\Downloads\se-propane-market-map\data\county_centroids.json'
if os.path.exists(centroid_file):
    with open(centroid_file) as f:
        centroids = json.load(f)
    print(f"Loaded {len(centroids)} centroids from cache")
else:
    print("\nNeed to download county centroids from US Census Gazetteer.")
    print("Will fetch from: https://www2.census.gov/geo/docs/maps-data/data/gazetteer/2023_Gazetteer/2023_Gaz_counties_national.txt")
    centroids = None

# ─── Step 6: Build national dataset with what we have ───
national = []
needs_coords = 0

for excel_name, vals in excel_counties.items():
    norm = normalize_name(excel_name)

    entry = {
        'n': norm,
        'h': vals['h'],
        'u': vals['u'],
        'g': vals['g'],
    }

    # Try to get coordinates from existing data
    if norm in existing_lookup:
        entry['lat'] = existing_lookup[norm]['lat']
        entry['lng'] = existing_lookup[norm]['lng']
    else:
        entry['lat'] = None
        entry['lng'] = None
        needs_coords += 1

    national.append(entry)

# Sort by total gallons descending
national.sort(key=lambda x: x['g'], reverse=True)

print(f"\nNational dataset: {len(national)} counties")
print(f"  With coordinates: {len(national) - needs_coords}")
print(f"  Need coordinates: {needs_coords}")

# Save intermediate result
with open(OUTPUT, 'w') as f:
    json.dump(national, f, separators=(',', ':'))
print(f"Saved to {OUTPUT}")

# Stats by state
state_stats = {}
for c in national:
    parts = c['n'].rsplit(', ', 1)
    if len(parts) == 2:
        st = parts[1]
        if st not in state_stats:
            state_stats[st] = {'count': 0, 'gallons': 0, 'has_coords': 0}
        state_stats[st]['count'] += 1
        state_stats[st]['gallons'] += c['g']
        if c['lat'] is not None:
            state_stats[st]['has_coords'] += 1

print(f"\n=== Top 20 States by Total Propane Consumption ===")
for st, s in sorted(state_stats.items(), key=lambda x: x[1]['gallons'], reverse=True)[:20]:
    print(f"  {st}: {s['count']} counties, {s['gallons']:,.0f} gal, {s['has_coords']}/{s['count']} coords")
